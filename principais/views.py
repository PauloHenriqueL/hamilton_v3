import csv
import json
import io
from rest_framework import generics
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, DetailView, UpdateView, DeleteView
from . import models, forms, serializers
from django.shortcuts import render, redirect
from rest_framework.permissions import IsAuthenticated
from app.permissions import GlobalDefaultPermission
from django.db.models import Q, Count, Prefetch, Avg, Sum
from django.http import JsonResponse, HttpResponse
from .models import Paciente, Terapeuta, Associado
from django.views.decorators.http import require_GET
from decimal import Decimal
from django.contrib import messages
from django.core.exceptions import ObjectDoesNotExist
from django.contrib.auth.decorators import login_required, permission_required
from django.utils import timezone
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from acessorios.models import Captacao



@login_required
@permission_required('principais.view_consulta')
def gerar_relatorio(request):
    """
    Gera relatórios em CSV ou Excel baseado nos parâmetros da requisição
    """
    tipo_relatorio = request.GET.get('tipo')
    formato = request.GET.get('formato', 'csv')  # csv ou excel
    
    if not tipo_relatorio:
        return HttpResponse('Tipo de relatório não especificado', status=400)
    
    # Definir nome do arquivo e extensão
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    if formato == 'excel':
        filename = f'{tipo_relatorio}_{timestamp}.xlsx'
        content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    else:  # CSV
        filename = f'{tipo_relatorio}_{timestamp}.csv'
        content_type = 'text/csv; charset=utf-8'
    
    # Preparar response
    response = HttpResponse(content_type=content_type)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    if formato == 'csv':
        response.write('\ufeff')  # BOM para UTF-8
        writer = csv.writer(response)
    else:  # Excel
        writer = None
    
    # Gerar relatório baseado no tipo
    if tipo_relatorio == 'associado':
        return _gerar_relatorio_associado(response, writer, formato)
    elif tipo_relatorio == 'paciente':
        return _gerar_relatorio_paciente(response, writer, formato)
    elif tipo_relatorio == 'terapeuta':
        return _gerar_relatorio_terapeuta(response, writer, formato)
    elif tipo_relatorio == 'avaliacao':
        return _gerar_relatorio_avaliacao(response, writer, formato)
    elif tipo_relatorio == 'consulta':
        return _gerar_relatorio_consulta(response, writer, formato)
    elif tipo_relatorio == 'dashboard':
        return _gerar_relatorio_dashboard(response, writer, formato)
    elif tipo_relatorio == 'metricas_terapeuta':
        return _gerar_relatorio_metricas_terapeuta(response, writer, formato)
    elif tipo_relatorio == 'altadesistencia':
        return _gerar_relatorio_altadesistencia(response, writer, formato)
    elif tipo_relatorio == 'selecao':
        return _gerar_relatorio_selecao(response, writer, formato)
    elif tipo_relatorio == 'captacao':
        return _gerar_relatorio_captacao(response, writer, formato)
    else:
        return HttpResponse('Tipo de relatório inválido', status=400)


def _criar_workbook_formatado():
    """Cria um workbook Excel com formatação padrão"""
    wb = Workbook()
    ws = wb.active
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    return wb, ws, header_font, header_fill, border


def _finalizar_excel(wb, response):
    """Finaliza e salva o workbook Excel no response"""
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response.write(buffer.getvalue())
    buffer.close()
    return response


def _gerar_relatorio_associado(response, writer, formato):
    """Gera relatório de Associados"""
    associados = Associado.objects.all().select_related().prefetch_related('setores')
    
    if formato == 'csv':
        writer.writerow(['ID', 'Nome', 'Email', 'Telefone', 'CPF', 'Endereco', 'Sexo', 'Data Nascimento', 'Setores', 'Ativo', 'Data Criacao'])
        for assoc in associados:
            setores = ', '.join([setor.setor for setor in assoc.setores.all()])
            writer.writerow([
                assoc.pk_associado, assoc.nome, assoc.email or '', assoc.telefone,
                assoc.cpf or '', assoc.endereco, assoc.sexo, 
                assoc.dat_nascimento or '', setores, assoc.is_active, assoc.created_at
            ])
    else:  # Excel
        wb, ws, header_font, header_fill, border = _criar_workbook_formatado()
        ws.title = "Associados"
        
        # Cabeçalhos
        headers = ['ID', 'Nome', 'Email', 'Telefone', 'CPF', 'Endereço', 'Sexo', 'Data Nascimento', 'Setores', 'Ativo', 'Data Criação']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Dados
        for row, assoc in enumerate(associados, 2):
            setores = ', '.join([setor.setor for setor in assoc.setores.all()])
            data = [
                assoc.pk_associado, assoc.nome, assoc.email or '', assoc.telefone,
                assoc.cpf or '', assoc.endereco, assoc.sexo, 
                assoc.dat_nascimento or '', setores, 'Sim' if assoc.is_active else 'Não', 
                assoc.created_at.strftime('%d/%m/%Y %H:%M')
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
        
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        return _finalizar_excel(wb, response)
    
    return response


def _gerar_relatorio_altadesistencia(response, writer, formato):
    """Gera relatório de Alta/Desistência"""
    altadesistencias = models.Altadesistencia.objects.all().select_related(
        'fk_terapeuta__fk_associado', 'fk_paciente'
    )
    
    if formato == 'csv':
        writer.writerow([
            'ID', 'Terapeuta', 'Paciente', 'Data Sessão', 'Cancelador', 
            'Motivo Cancelamento', 'Momento', 'Alta/Desistência', 'Data Criação'
        ])
        for alta in altadesistencias:
            writer.writerow([
                alta.pk_alta_desistencia,
                alta.fk_terapeuta.fk_associado.nome,
                alta.fk_paciente.nome,
                alta.dat_sessao or '',
                alta.cancelador or '',
                alta.motivo_cancel or '',
                alta.momento or '',
                alta.alta_desistencia or '',
                alta.created_at
            ])
    else:  # Excel
        wb, ws, header_font, header_fill, border = _criar_workbook_formatado()
        ws.title = "Alta_Desistencia"
        
        # Cabeçalhos
        headers = [
            'ID', 'Terapeuta', 'Paciente', 'Data Sessão', 'Cancelador', 
            'Motivo Cancelamento', 'Momento', 'Alta/Desistência', 'Data Criação'
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Dados
        for row, alta in enumerate(altadesistencias, 2):
            data = [
                alta.pk_alta_desistencia,
                alta.fk_terapeuta.fk_associado.nome,
                alta.fk_paciente.nome,
                alta.dat_sessao.strftime('%d/%m/%Y') if alta.dat_sessao else '',
                alta.cancelador or '',
                alta.motivo_cancel or '',
                alta.momento or '',
                alta.alta_desistencia or '',
                alta.created_at.strftime('%d/%m/%Y %H:%M')
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
        
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        return _finalizar_excel(wb, response)
    
    return response


def _gerar_relatorio_selecao(response, writer, formato):
    """Gera relatório de Seleções (Avaliações de Terapeutas)"""
    selecoes = models.Selecao.objects.all().select_related(
        'fk_terapeuta_avaliador__fk_associado', 'fk_associado_avaliado'
    )
    
    if formato == 'csv':
        writer.writerow([
            'ID', 'Avaliador', 'Avaliado', 'Data Avaliação', 'Estágio Mudança',
            'Estrutura', 'Encerramento', 'Acolhimento', 'Segurança Terapeuta',
            'Segurança Método', 'Aprofundar', 'Hipóteses', 'Interpretação',
            'Frase & Timing', 'Corpo & Setting', 'Insight & Potência',
            'Média Geral', 'Data Criação'
        ])
        for selecao in selecoes:
            # Calcular média geral das avaliações
            campos_avaliacao = [
                selecao.estagio_mudanca, selecao.estrutura, selecao.encerramento,
                selecao.acolhimento, selecao.seguranca_terapeuta, selecao.seguranca_metodo,
                selecao.aprofundar, selecao.hipoteses, selecao.interpretacao,
                selecao.frase_timing, selecao.corpo_setting, selecao.insight_potencia
            ]
            media_geral = sum(campos_avaliacao) / len(campos_avaliacao)
            
            writer.writerow([
                selecao.pk_selecao,
                selecao.fk_terapeuta_avaliador.fk_associado.nome,
                selecao.fk_associado_avaliado.nome,
                selecao.dat_avaliacao,
                selecao.estagio_mudanca,
                selecao.estrutura,
                selecao.encerramento,
                selecao.acolhimento,
                selecao.seguranca_terapeuta,
                selecao.seguranca_metodo,
                selecao.aprofundar,
                selecao.hipoteses,
                selecao.interpretacao,
                selecao.frase_timing,
                selecao.corpo_setting,
                selecao.insight_potencia,
                f'{media_geral:.2f}',
                selecao.created_at if hasattr(selecao, 'created_at') else ''
            ])
    else:  # Excel
        wb, ws, header_font, header_fill, border = _criar_workbook_formatado()
        ws.title = "Selecoes"
        
        # Cabeçalhos
        headers = [
            'ID', 'Avaliador', 'Avaliado', 'Data Avaliação', 'Estágio Mudança',
            'Estrutura', 'Encerramento', 'Acolhimento', 'Segurança Terapeuta',
            'Segurança Método', 'Aprofundar', 'Hipóteses', 'Interpretação',
            'Frase & Timing', 'Corpo & Setting', 'Insight & Potência',
            'Média Geral'
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Dados
        for row, selecao in enumerate(selecoes, 2):
            # Calcular média geral das avaliações
            campos_avaliacao = [
                selecao.estagio_mudanca, selecao.estrutura, selecao.encerramento,
                selecao.acolhimento, selecao.seguranca_terapeuta, selecao.seguranca_metodo,
                selecao.aprofundar, selecao.hipoteses, selecao.interpretacao,
                selecao.frase_timing, selecao.corpo_setting, selecao.insight_potencia
            ]
            media_geral = sum(campos_avaliacao) / len(campos_avaliacao)
            
            data = [
                selecao.pk_selecao,
                selecao.fk_terapeuta_avaliador.fk_associado.nome,
                selecao.fk_associado_avaliado.nome,
                selecao.dat_avaliacao.strftime('%d/%m/%Y'),
                selecao.estagio_mudanca,
                selecao.estrutura,
                selecao.encerramento,
                selecao.acolhimento,
                selecao.seguranca_terapeuta,
                selecao.seguranca_metodo,
                selecao.aprofundar,
                selecao.hipoteses,
                selecao.interpretacao,
                selecao.frase_timing,
                selecao.corpo_setting,
                selecao.insight_potencia,
                f'{media_geral:.2f}'
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
        
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        return _finalizar_excel(wb, response)
    
    return response


def _gerar_relatorio_captacao(response, writer, formato):
    """Gera relatório de Captações com quantidade de pacientes por captação"""
    # Buscar captações com contagem de pacientes
    captacoes_com_contagem = Captacao.objects.annotate(
        total_pacientes=Count('paciente', filter=Q(paciente__is_active=True)),
        total_pacientes_inativos=Count('paciente', filter=Q(paciente__is_active=False)),
        total_geral=Count('paciente')
    ).order_by('-total_pacientes')
    
    if formato == 'csv':
        writer.writerow([
            'ID', 'Nome da Captação', 'Pacientes Ativos', 'Pacientes Inativos', 
            'Total de Pacientes', 'Ativo', 'Data Criação'
        ])
        for captacao in captacoes_com_contagem:
            writer.writerow([
                captacao.pk_captacao,
                captacao.nome,
                captacao.total_pacientes,
                captacao.total_pacientes_inativos,
                captacao.total_geral,
                'Sim' if captacao.is_active else 'Não',
                captacao.created_at
            ])
            
        # Adicionar estatísticas resumo
        writer.writerow([])  # Linha vazia
        writer.writerow(['=== ESTATÍSTICAS RESUMO ==='])
        
        total_captacoes_ativas = captacoes_com_contagem.filter(is_active=True).count()
        total_captacoes_inativas = captacoes_com_contagem.filter(is_active=False).count()
        captacao_mais_usada = captacoes_com_contagem.first()
        
        writer.writerow(['Total de Captações Ativas', total_captacoes_ativas])
        writer.writerow(['Total de Captações Inativas', total_captacoes_inativas])
        if captacao_mais_usada:
            writer.writerow(['Captação Mais Utilizada', f'{captacao_mais_usada.nome} ({captacao_mais_usada.total_pacientes} pacientes)'])
            
    else:  # Excel
        wb, ws, header_font, header_fill, border = _criar_workbook_formatado()
        ws.title = "Captacoes"
        
        # Cabeçalhos
        headers = [
            'ID', 'Nome da Captação', 'Pacientes Ativos', 'Pacientes Inativos', 
            'Total de Pacientes', 'Ativo', 'Data Criação'
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Dados
        row_num = 2
        for captacao in captacoes_com_contagem:
            data = [
                captacao.pk_captacao,
                captacao.nome,
                captacao.total_pacientes,
                captacao.total_pacientes_inativos,
                captacao.total_geral,
                'Sim' if captacao.is_active else 'Não',
                captacao.created_at.strftime('%d/%m/%Y %H:%M')
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = border
            row_num += 1
        
        # Adicionar estatísticas resumo
        row_num += 2  # Pular uma linha
        resumo_header = ws.cell(row=row_num, column=1, value="ESTATÍSTICAS RESUMO")
        resumo_header.font = header_font
        resumo_header.fill = header_fill
        row_num += 1
        
        total_captacoes_ativas = captacoes_com_contagem.filter(is_active=True).count()
        total_captacoes_inativas = captacoes_com_contagem.filter(is_active=False).count()
        captacao_mais_usada = captacoes_com_contagem.first()
        
        estatisticas = [
            ['Total de Captações Ativas', total_captacoes_ativas],
            ['Total de Captações Inativas', total_captacoes_inativas],
        ]
        
        if captacao_mais_usada:
            estatisticas.append(['Captação Mais Utilizada', f'{captacao_mais_usada.nome} ({captacao_mais_usada.total_pacientes} pacientes)'])
        
        for stat_nome, stat_valor in estatisticas:
            ws.cell(row=row_num, column=1, value=stat_nome).border = border
            ws.cell(row=row_num, column=2, value=stat_valor).border = border
            row_num += 1
        
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        return _finalizar_excel(wb, response)
    
    return response


def _gerar_relatorio_paciente(response, writer, formato):

    pacientes = Paciente.objects.all().select_related('fk_clinica', 'fk_modalidade', 'fk_captacao')
    
    if formato == 'csv':
        writer.writerow(['ID', 'Nome', 'Email', 'Telefone', 'Clinica', 'Modalidade', 'Captacao', 'Valor Sessao', 'Data Nascimento', 'Ativo', 'Data Criacao'])
        for pac in pacientes:
            writer.writerow([
                pac.pk_paciente, pac.nome, pac.email or '', pac.telefone,
                pac.fk_clinica.clinica, pac.fk_modalidade.modalidade, pac.fk_captacao.nome,
                pac.vlr_sessao, pac.dat_nascimento or '', pac.is_active, pac.created_at
            ])
    else:  # Excel
        wb, ws, header_font, header_fill, border = _criar_workbook_formatado()
        ws.title = "Pacientes"
        
        # Cabeçalhos
        headers = ['ID', 'Nome', 'Email', 'Telefone', 'Clínica', 'Modalidade', 'Captação', 'Valor Sessão', 'Data Nascimento', 'Ativo', 'Data Criação']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Dados
        for row, pac in enumerate(pacientes, 2):
            data = [
                pac.pk_paciente, pac.nome, pac.email or '', pac.telefone,
                pac.fk_clinica.clinica, pac.fk_modalidade.modalidade, pac.fk_captacao.nome,
                float(pac.vlr_sessao), pac.dat_nascimento or '', 'Sim' if pac.is_active else 'Não',
                pac.created_at.strftime('%d/%m/%Y %H:%M')
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
        
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        return _finalizar_excel(wb, response)
    
    return response


def _gerar_relatorio_terapeuta(response, writer, formato):
    """Gera relatório de Terapeutas"""
    terapeutas = Terapeuta.objects.all().select_related(
        'fk_associado', 'fk_decano', 'fk_abordagem', 'fk_nucleo', 'fk_clinica', 'fk_modalidade'
    )
    
    if formato == 'csv':
        writer.writerow(['ID', 'Nome', 'Decano', 'Abordagem', 'Nucleo', 'Clinica', 'Modalidade', 'Ativo', 'Data Criacao'])
        for ter in terapeutas:
            writer.writerow([
                ter.pk_terapeuta, ter.fk_associado.nome, ter.fk_decano.nome,
                ter.fk_abordagem.abordagem, ter.fk_nucleo.nucleo, ter.fk_clinica.clinica,
                ter.fk_modalidade.modalidade, ter.is_active, ter.created_at
            ])
    else:  # Excel
        wb, ws, header_font, header_fill, border = _criar_workbook_formatado()
        ws.title = "Terapeutas"
        
        # Cabeçalhos
        headers = ['ID', 'Nome', 'Decano', 'Abordagem', 'Núcleo', 'Clínica', 'Modalidade', 'Ativo', 'Data Criação']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Dados
        for row, ter in enumerate(terapeutas, 2):
            data = [
                ter.pk_terapeuta, ter.fk_associado.nome, ter.fk_decano.nome,
                ter.fk_abordagem.abordagem, ter.fk_nucleo.nucleo, ter.fk_clinica.clinica,
                ter.fk_modalidade.modalidade, 'Sim' if ter.is_active else 'Não',
                ter.created_at.strftime('%d/%m/%Y %H:%M')
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
        
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        return _finalizar_excel(wb, response)
    
    return response


def _gerar_relatorio_avaliacao(response, writer, formato):
    """Gera relatório de Avaliações"""
    avaliacoes = models.Avaliacao.objects.all().select_related('fk_terapeuta__fk_associado', 'fk_paciente')
    
    if formato == 'csv':
        writer.writerow(['ID', 'Terapeuta', 'Paciente', 'Data Consulta', 'Individual', 'Interpessoal', 'Social', 'Geral', 'Qualidade Geral', 'Momento', 'Data Criacao'])
        for av in avaliacoes:
            writer.writerow([
                av.pk_avaliacao, av.fk_terapeuta.fk_associado.nome, av.fk_paciente.nome,
                av.dat_consulta, av.individual or '', av.interpessoal or '', av.social or '',
                av.geral or '', av.qualidade_geral or '', av.momento, av.created_at
            ])
    else:  # Excel
        wb, ws, header_font, header_fill, border = _criar_workbook_formatado()
        ws.title = "Avaliações"
        
        # Cabeçalhos
        headers = ['ID', 'Terapeuta', 'Paciente', 'Data Consulta', 'Individual', 'Interpessoal', 'Social', 'Geral', 'Qualidade Geral', 'Momento', 'Data Criação']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Dados
        for row, av in enumerate(avaliacoes, 2):
            data = [
                av.pk_avaliacao, av.fk_terapeuta.fk_associado.nome, av.fk_paciente.nome,
                av.dat_consulta.strftime('%d/%m/%Y'), av.individual or '', av.interpessoal or '', av.social or '',
                av.geral or '', av.qualidade_geral or '', av.momento, 
                av.created_at.strftime('%d/%m/%Y %H:%M')
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
        
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        return _finalizar_excel(wb, response)
    
    return response


def _gerar_relatorio_consulta(response, writer, formato):
    """Gera relatório de Consultas"""
    consultas = models.Consulta.objects.all().select_related('fk_terapeuta__fk_associado', 'fk_paciente')
    
    if formato == 'csv':
        writer.writerow(['ID', 'Terapeuta', 'Paciente', 'Data Consulta', 'Valor Consulta', 'Valor Pago', 'Realizada', 'Data Criacao'])
        for cons in consultas:
            writer.writerow([
                cons.pk_consulta, cons.fk_terapeuta.fk_associado.nome, cons.fk_paciente.nome,
                cons.dat_consulta, cons.vlr_consulta, cons.vlr_pago or '', cons.is_realizado, cons.created_at
            ])
    else:  # Excel
        wb, ws, header_font, header_fill, border = _criar_workbook_formatado()
        ws.title = "Consultas"
        
        # Cabeçalhos
        headers = ['ID', 'Terapeuta', 'Paciente', 'Data Consulta', 'Valor Consulta', 'Valor Pago', 'Realizada', 'Data Criação']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Dados
        for row, cons in enumerate(consultas, 2):
            data = [
                cons.pk_consulta, cons.fk_terapeuta.fk_associado.nome, cons.fk_paciente.nome,
                cons.dat_consulta.strftime('%d/%m/%Y'), float(cons.vlr_consulta), 
                float(cons.vlr_pago) if cons.vlr_pago else '', 'Sim' if cons.is_realizado else 'Não',
                cons.created_at.strftime('%d/%m/%Y %H:%M')
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
        
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        return _finalizar_excel(wb, response)
    
    return response


def _gerar_relatorio_dashboard(response, writer, formato):
    """Gera relatório com métricas do dashboard"""
    # Calcular métricas (você precisará adaptar essa lógica baseada no seu dashboard atual)
    total_consultas_marcadas = models.Consulta.objects.count()
    total_consultas_realizadas = models.Consulta.objects.filter(is_realizado=True).count()
    taxa_adesao = (total_consultas_realizadas / total_consultas_marcadas * 100) if total_consultas_marcadas > 0 else 0
    
    receita_total = models.Consulta.objects.filter(vlr_pago__isnull=False).aggregate(total=Sum('vlr_pago'))['total'] or 0
    pacientes_ativos = models.Paciente.objects.filter(is_active=True).count()
    terapeutas_ativos = models.Terapeuta.objects.filter(is_active=True).count()
    
    if formato == 'csv':
        writer.writerow(['Metrica', 'Valor'])
        writer.writerow(['Total Consultas Marcadas', total_consultas_marcadas])
        writer.writerow(['Total Consultas Realizadas', total_consultas_realizadas])
        writer.writerow(['Taxa de Adesao (%)', f'{taxa_adesao:.2f}'])
        writer.writerow(['Receita Total Recebida', f'{receita_total:.2f}'])
        writer.writerow(['Pacientes Ativos', pacientes_ativos])
        writer.writerow(['Terapeutas Ativos', terapeutas_ativos])
        writer.writerow(['Data Geracao', timezone.now().strftime('%Y-%m-%d %H:%M:%S')])
    else:  # Excel
        wb, ws, header_font, header_fill, border = _criar_workbook_formatado()
        ws.title = "Dashboard Métricas"
        
        # Cabeçalhos
        headers = ['Métrica', 'Valor']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Dados
        metricas = [
            ['Total Consultas Marcadas', total_consultas_marcadas],
            ['Total Consultas Realizadas', total_consultas_realizadas],
            ['Taxa de Adesão (%)', f'{taxa_adesao:.2f}%'],
            ['Receita Total Recebida', f'R$ {receita_total:.2f}'],
            ['Pacientes Ativos', pacientes_ativos],
            ['Terapeutas Ativos', terapeutas_ativos],
            ['Data Geração', timezone.now().strftime('%d/%m/%Y %H:%M:%S')]
        ]
        
        for row, (metrica, valor) in enumerate(metricas, 2):
            ws.cell(row=row, column=1, value=metrica).border = border
            ws.cell(row=row, column=2, value=valor).border = border
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        
        return _finalizar_excel(wb, response)
    
    return response


def _gerar_relatorio_metricas_terapeuta(response, writer, formato):
    """Gera relatório de métricas por terapeuta"""
    terapeutas = models.Terapeuta.objects.filter(is_active=True).select_related('fk_associado')
    
    if formato == 'csv':
        writer.writerow(['Terapeuta', 'Pacientes Ativos', 'Consultas Marcadas', 'Consultas Realizadas', 'Taxa Adesao (%)', 'Valor Recebido', 'Receita Acordada'])
        
        for terapeuta in terapeutas:
            # Calcular métricas por terapeuta
            consultas_marcadas = models.Consulta.objects.filter(fk_terapeuta=terapeuta).count()
            consultas_realizadas = models.Consulta.objects.filter(fk_terapeuta=terapeuta, is_realizado=True).count()
            taxa_adesao = (consultas_realizadas / consultas_marcadas * 100) if consultas_marcadas > 0 else 0
            
            valor_recebido = models.Consulta.objects.filter(
                fk_terapeuta=terapeuta, 
                vlr_pago__isnull=False
            ).aggregate(total=Sum('vlr_pago'))['total'] or 0
            
            pacientes_ativos = models.Consulta.objects.filter(
                fk_terapeuta=terapeuta
            ).values('fk_paciente').distinct().count()
            
            # Receita acordada (você pode precisar ajustar essa lógica)
            receita_acordada = models.Paciente.objects.filter(
                consulta__fk_terapeuta=terapeuta
            ).aggregate(total=Sum('vlr_sessao'))['total'] or 0
            
            writer.writerow([
                terapeuta.fk_associado.nome,
                pacientes_ativos,
                consultas_marcadas,
                consultas_realizadas,
                f'{taxa_adesao:.2f}',
                f'{valor_recebido:.2f}',
                f'{receita_acordada:.2f}'
            ])
    else:  # Excel
        wb, ws, header_font, header_fill, border = _criar_workbook_formatado()
        ws.title = "Métricas por Terapeuta"
        
        # Cabeçalhos
        headers = ['Terapeuta', 'Pacientes Ativos', 'Consultas Marcadas', 'Consultas Realizadas', 'Taxa Adesão (%)', 'Valor Recebido (R$)', 'Receita Acordada (R$)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Dados
        for row, terapeuta in enumerate(terapeutas, 2):
            # Calcular métricas por terapeuta (mesmo cálculo do CSV)
            consultas_marcadas = models.Consulta.objects.filter(fk_terapeuta=terapeuta).count()
            consultas_realizadas = models.Consulta.objects.filter(fk_terapeuta=terapeuta, is_realizado=True).count()
            taxa_adesao = (consultas_realizadas / consultas_marcadas * 100) if consultas_marcadas > 0 else 0
            
            valor_recebido = models.Consulta.objects.filter(
                fk_terapeuta=terapeuta, 
                vlr_pago__isnull=False
            ).aggregate(total=Sum('vlr_pago'))['total'] or 0
            
            pacientes_ativos = models.Consulta.objects.filter(
                fk_terapeuta=terapeuta
            ).values('fk_paciente').distinct().count()
            
            receita_acordada = models.Paciente.objects.filter(
                consulta__fk_terapeuta=terapeuta
            ).aggregate(total=Sum('vlr_sessao'))['total'] or 0
            
            data = [
                terapeuta.fk_associado.nome,
                pacientes_ativos,
                consultas_marcadas,
                consultas_realizadas,
                f'{taxa_adesao:.2f}%',
                f'R$ {valor_recebido:.2f}',
                f'R$ {receita_acordada:.2f}'
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
        
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        return _finalizar_excel(wb, response)
    
    return response

@require_GET
def paciente_valor_sessao(request, pk):
    """Endpoint para obter o valor da sessão de um paciente"""
    try:
        paciente = Paciente.objects.get(pk=pk)
        return JsonResponse({'vlr_sessao': int(paciente.vlr_sessao)})
    except Paciente.DoesNotExist:
        return JsonResponse({'error': 'Paciente não encontrado'}, status=404)


class ConsultaListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = models.Consulta
    template_name = 'consulta_list.html'
    context_object_name = 'consultas'
    paginate_by = 10
    permission_required = 'principais.view_consulta'

    def get_queryset(self):

        queryset = models.Consulta.objects.select_related(
            'fk_terapeuta',
            'fk_paciente',
            'fk_terapeuta__fk_associado',
            'fk_terapeuta__fk_abordagem',
            'fk_terapeuta__fk_clinica'
        )
        
        try:
            associado = Associado.objects.get(usuario=self.request.user)
            terapeuta = Terapeuta.objects.get(fk_associado=associado)
            queryset = queryset.filter(fk_terapeuta=terapeuta)
        except (Associado.DoesNotExist, Terapeuta.DoesNotExist):
            if not self.request.user.is_staff:
                queryset = queryset.none()
        
        nome = self.request.GET.get('nome')
        if nome:
            queryset = queryset.filter(
                Q(fk_paciente__nome__icontains=nome) | 
                Q(fk_terapeuta__fk_associado__nome__icontains=nome)
            )
        
        order_by = self.request.GET.get('order_by', '-dat_consulta')

        valid_orders = [
            'pk_consulta', '-pk_consulta',
            'fk_terapeuta__fk_associado__nome', '-fk_terapeuta__fk_associado__nome',
            'fk_paciente__nome', '-fk_paciente__nome',
            'dat_consulta', '-dat_consulta',
            'is_realizado', '-is_realizado',
            'vlr_pago', '-vlr_pago'
        ]
        
        if order_by in valid_orders:
            queryset = queryset.order_by(order_by)
        else:
            queryset = queryset.order_by('-dat_consulta')
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['current_order'] = self.request.GET.get('order_by', '-dat_consulta')
        context['current_search'] = self.request.GET.get('nome', '')
        return context

class ConsultaCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = models.Consulta
    template_name = 'consulta_create.html'
    form_class = forms.ConsultaForm
    success_url = reverse_lazy('consulta-list')
    permission_required = 'principais.add_consulta'
    
    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        
        # Remover campos que não serão usados no formulário
        for field_name in ['vlr_consulta', 'vlr_pago']:
            if field_name in form.fields:
                del form.fields[field_name]
        
        try:
            # Buscar o associado do usuário logado
            associado = Associado.objects.get(usuario=self.request.user)
            # Buscar o terapeuta relacionado ao associado
            terapeuta = Terapeuta.objects.get(fk_associado=associado)
            
            form.fields['fk_terapeuta'].initial = terapeuta.pk_terapeuta
            form.fields['fk_terapeuta'].widget.attrs.update({
                'readonly': 'readonly',
                'style': 'pointer-events: none; background-color: #343a40 !important; color: #ffffff !important; border-color: #495057 !important;'
            })
            form.fields['fk_terapeuta'].queryset = form.fields['fk_terapeuta'].queryset.filter(pk=terapeuta.pk_terapeuta)
            
        except (Associado.DoesNotExist, Terapeuta.DoesNotExist):
            # Se não encontrar terapeuta, permitir seleção livre (para staff)
            pass
        
        return form
    
    def form_valid(self, form):
        quantidade = int(self.request.POST.get('quantidade', 1))
        vlr_pix_total_str = self.request.POST.get('vlr_pix_total', '')
        
        # Validar se o valor PIX foi preenchido
        if not vlr_pix_total_str or not vlr_pix_total_str.strip():
            form.add_error(None, 'O valor total recebido no PIX é obrigatório.')
            return self.form_invalid(form)
        
        vlr_pix_total = float(vlr_pix_total_str)
        
        # Calcular valor por consulta (PIX dividido pela quantidade)
        vlr_consulta = round(vlr_pix_total / quantidade, 2) if quantidade > 0 else 0
        
        # Validar se todas as datas foram preenchidas
        for i in range(quantidade):
            data_key = f'data_consulta_{i}'
            if not self.request.POST.get(data_key):
                form.add_error(None, f'Data da consulta {i+1} é obrigatória.')
                return self.form_invalid(form)
        
        consulta = form.save(commit=False)
        
        # Definir terapeuta se o usuário for um terapeuta
        try:
            associado = Associado.objects.get(usuario=self.request.user)
            terapeuta = Terapeuta.objects.get(fk_associado=associado)
            consulta.fk_terapeuta = terapeuta
        except (Associado.DoesNotExist, Terapeuta.DoesNotExist):
            # Se não for terapeuta, usar o terapeuta selecionado no form
            if not consulta.fk_terapeuta:
                messages.error(self.request, 'Erro: Terapeuta não foi definido corretamente.')
                return self.form_invalid(form)
        
        # Primeira consulta
        consulta.dat_consulta = self.request.POST.get('data_consulta_0')
        consulta.vlr_consulta = vlr_consulta  # Valor calculado
        consulta.vlr_pago = vlr_consulta      # Mesmo valor (pago via PIX)
        
        # Capturar valores dos checkboxes do formulário (primeira consulta)
        is_realizado_0 = self.request.POST.get('is_realizado_0') == 'on'
        consulta.is_realizado = is_realizado_0
        
        consulta.save()
        
        # Consultas adicionais
        if quantidade > 1:
            for i in range(1, quantidade):
                nova_consulta = models.Consulta(
                    fk_terapeuta=consulta.fk_terapeuta,
                    fk_paciente=consulta.fk_paciente,
                    vlr_consulta=vlr_consulta,  # Valor calculado
                    vlr_pago=vlr_consulta,      # Mesmo valor (pago via PIX)
                )
                
                nova_consulta.dat_consulta = self.request.POST.get(f'data_consulta_{i}')
                
                # Capturar valores dos checkboxes específicos de cada consulta
                is_realizado_i = self.request.POST.get(f'is_realizado_{i}') == 'on'
                nova_consulta.is_realizado = is_realizado_i
                
                nova_consulta.save()
        
        messages.success(self.request, f'{quantidade} consulta(s) cadastrada(s) com sucesso! Valor por consulta: R$ {vlr_consulta:.2f}')
        return redirect(self.success_url)


class ConsultaDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = models.Consulta
    template_name = 'consulta_detail.html'
    permission_required = 'principais.view_consulta'
    
    def get_queryset(self):
        # Otimização: Buscar dados relacionados junto
        return models.Consulta.objects.select_related(
            'fk_terapeuta',
            'fk_paciente',
            'fk_terapeuta__fk_associado',
            'fk_terapeuta__fk_abordagem',
            'fk_terapeuta__fk_clinica',
            'fk_terapeuta__fk_decano',
            'fk_terapeuta__fk_nucleo',
            'fk_terapeuta__fk_modalidade',
            'fk_paciente__fk_clinica',
            'fk_paciente__fk_modalidade'
        )
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        consulta = self.object
        
        if consulta.vlr_pago is not None and consulta.vlr_consulta is not None:
            context['diferenca_valor'] = consulta.vlr_pago - consulta.vlr_consulta
        else:
            context['diferenca_valor'] = Decimal('0.00')
            
        return context


class ConsultaUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = models.Consulta
    template_name = 'consulta_update.html'
    form_class = forms.ConsultaForm
    permission_required = 'principais.change_consulta'
    
    def get_success_url(self):
        return reverse_lazy('consulta-detail', kwargs={'pk': self.object.pk})
    
    def get_queryset(self):
        # Otimização: Buscar dados relacionados junto
        return models.Consulta.objects.select_related(
            'fk_terapeuta',
            'fk_paciente'
        )
    
    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        # Remover campos que não são usados na edição
        for field_name in ['quantidade', 'vlr_pix_total']:
            if field_name in form.fields:
                del form.fields[field_name]
        return form
    
    def form_valid(self, form):
        consulta = form.save(commit=False)
        
        # Capturar dados adicionais do formulário
        consulta.dat_consulta = self.request.POST.get('dat_consulta')
        
        # Capturar checkboxes
        is_realizado = self.request.POST.get('is_realizado') == 'on'
        is_pago = self.request.POST.get('is_pago') == 'on'
        
        consulta.is_realizado = is_realizado
        
        # Se foi marcado como pago, usar o valor da consulta se vlr_pago não foi definido
        if is_pago and not consulta.vlr_pago:
            consulta.vlr_pago = consulta.vlr_consulta
        elif not is_pago:
            consulta.vlr_pago = 0
        
        consulta.save()
        messages.success(self.request, 'Consulta atualizada com sucesso!')
        return redirect(self.get_success_url())


class ConsultaDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = models.Consulta
    template_name = 'consulta_delete.html'
    success_url = reverse_lazy('consulta-list')
    permission_required = 'principais.delete_consulta'
    
    def get_queryset(self):
        # Otimização: Buscar dados relacionados para mostrar no template
        return models.Consulta.objects.select_related(
            'fk_terapeuta',
            'fk_paciente',
            'fk_terapeuta__fk_associado'
        )
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Consulta excluída com sucesso!')
        return super().delete(request, *args, **kwargs)


class MatchCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = models.Match
    template_name = 'match_create.html'
    form_class = forms.MatchForm
    success_url = reverse_lazy('consulta-list')
    permission_required = 'principais.add_match'
    
    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        
        try:
            # Buscar o associado do usuário logado
            associado = Associado.objects.get(usuario=self.request.user)
            # Buscar o terapeuta relacionado ao associado
            terapeuta = Terapeuta.objects.get(fk_associado=associado)
            
            form.fields['fk_terapeuta'].initial = terapeuta.pk_terapeuta
            form.fields['fk_terapeuta'].widget.attrs.update({
                'readonly': 'readonly',
                'style': 'pointer-events: none; background-color: #343a40 !important; color: #ffffff !important; border-color: #495057 !important;'
            })
            form.fields['fk_terapeuta'].queryset = form.fields['fk_terapeuta'].queryset.filter(pk=terapeuta.pk_terapeuta)
            
        except (Associado.DoesNotExist, Terapeuta.DoesNotExist):
            # Se não encontrar terapeuta, permitir seleção livre (para staff)
            pass
        
        return form
    
    def form_valid(self, form):
        match = form.save(commit=False)
        
        try:
            associado = Associado.objects.get(usuario=self.request.user)
            terapeuta = Terapeuta.objects.get(fk_associado=associado)
            match.fk_terapeuta = terapeuta
        except (Associado.DoesNotExist, Terapeuta.DoesNotExist):
            if not match.fk_terapeuta:
                messages.error(self.request, 'Erro: Terapeuta não foi definido corretamente.')
                return self.form_invalid(form)
        
        match.save()
        messages.success(self.request, 'Match cadastrada com sucesso!')
        
        return redirect(self.success_url)


class AltaDesistenciaCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = models.Altadesistencia
    template_name = 'altadesistencia_create.html'
    form_class = forms.AltaDesistenciaForm
    success_url = reverse_lazy('consulta-list')
    permission_required = 'principais.add_altadesistencia'
    
    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        
        try:
            # Buscar o associado do usuário logado
            associado = Associado.objects.get(usuario=self.request.user)
            # Buscar o terapeuta relacionado ao associado
            terapeuta = Terapeuta.objects.get(fk_associado=associado)
            
            form.fields['fk_terapeuta'].initial = terapeuta.pk_terapeuta
            form.fields['fk_terapeuta'].widget.attrs.update({
                'readonly': 'readonly',
                'style': 'pointer-events: none; background-color: #343a40 !important; color: #ffffff !important; border-color: #495057 !important;'
            })
            form.fields['fk_terapeuta'].queryset = form.fields['fk_terapeuta'].queryset.filter(pk=terapeuta.pk_terapeuta)
            
        except (Associado.DoesNotExist, Terapeuta.DoesNotExist):
            # Se não encontrar terapeuta, permitir seleção livre (para staff)
            pass
        
        return form
    
    def form_valid(self, form):
        altadesistencia = form.save(commit=False)
        
        try:
            associado = Associado.objects.get(usuario=self.request.user)
            terapeuta = Terapeuta.objects.get(fk_associado=associado)
            altadesistencia.fk_terapeuta = terapeuta
        except (Associado.DoesNotExist, Terapeuta.DoesNotExist):
            if not altadesistencia.fk_terapeuta:
                messages.error(self.request, 'Erro: Terapeuta não foi definido corretamente.')
                return self.form_invalid(form)
        
        altadesistencia.save()
        messages.success(self.request, 'Alta/Desistência cadastrada com sucesso!')
        
        return redirect(self.success_url)


# API VIEWS OTIMIZADAS
class ConsultaListCreateAPIView(generics.ListCreateAPIView):
    serializer_class = serializers.ConsultaSerializer
    permission_classes = (IsAuthenticated, GlobalDefaultPermission,)
    
    def get_queryset(self):
        # Otimização: Buscar dados relacionados nas APIs
        return models.Consulta.objects.select_related(
            'fk_terapeuta',
            'fk_paciente',
            'fk_terapeuta__fk_associado',
            'fk_terapeuta__fk_abordagem',
            'fk_terapeuta__fk_clinica'
        )


class ConsultaRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = serializers.ConsultaSerializer
    permission_classes = (IsAuthenticated, GlobalDefaultPermission,)
    
    def get_queryset(self):
        return models.Consulta.objects.select_related(
            'fk_terapeuta',
            'fk_paciente',
            'fk_terapeuta__fk_associado',
            'fk_terapeuta__fk_abordagem',
            'fk_terapeuta__fk_clinica'
        )


class TerapeutaListCreateAPIView(generics.ListCreateAPIView):
    permission_classes = (IsAuthenticated, GlobalDefaultPermission,)
    serializer_class = serializers.TerapeutaSerializer
    
    def get_queryset(self):
        # Otimização: Buscar dados relacionados + estatísticas
        return models.Terapeuta.objects.select_related(
            'fk_associado',
            'fk_abordagem',
            'fk_clinica',
            'fk_decano'
        ).annotate(
            total_consultas=Count('consulta'),
            total_pacientes=Count('consulta__fk_paciente', distinct=True)
        )


class TerapeutaRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = (IsAuthenticated, GlobalDefaultPermission,)
    serializer_class = serializers.TerapeutaSerializer
    
    def get_queryset(self):
        return models.Terapeuta.objects.select_related(
            'fk_associado',
            'fk_abordagem',
            'fk_clinica',
            'fk_decano'
        ).annotate(
            total_consultas=Count('consulta'),
            total_pacientes=Count('consulta__fk_paciente', distinct=True)
        )


class PacienteListCreateAPIView(generics.ListCreateAPIView):
    permission_classes = (IsAuthenticated, GlobalDefaultPermission,)
    queryset = models.Paciente.objects.select_related('fk_clinica', 'fk_captacao', 'fk_modalidade')
    serializer_class = serializers.PacienteSerializer


class PacienteRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = (IsAuthenticated, GlobalDefaultPermission,)
    queryset = models.Paciente.objects.select_related('fk_clinica', 'fk_captacao', 'fk_modalidade')
    serializer_class = serializers.PacienteSerializer


class AvaliacaoListCreateAPIView(generics.ListCreateAPIView):
    permission_classes = (IsAuthenticated, GlobalDefaultPermission,)
    queryset = models.Avaliacao.objects.select_related('fk_terapeuta__fk_associado', 'fk_paciente')
    serializer_class = serializers.AvaliacaoSerializer


class AvaliacaoRetrieveUpdateDestroyAPIView(generics.RetrieveDestroyAPIView):
    permission_classes = (IsAuthenticated, GlobalDefaultPermission,)
    queryset = models.Avaliacao.objects.select_related('fk_terapeuta__fk_associado', 'fk_paciente')
    serializer_class = serializers.AvaliacaoSerializer


class AltadesistenciaListCreateAPIView(generics.ListCreateAPIView):
    permission_classes = (IsAuthenticated, GlobalDefaultPermission,)
    queryset = models.Altadesistencia.objects.select_related('fk_terapeuta__fk_associado', 'fk_paciente')
    serializer_class = serializers.AltadesistenciaSerializer


class AltadesistenciaRetrieveUpdateDestroyAPIView(generics.RetrieveDestroyAPIView):
    permission_classes = (IsAuthenticated, GlobalDefaultPermission,)
    queryset = models.Altadesistencia.objects.select_related('fk_terapeuta__fk_associado', 'fk_paciente')
    serializer_class = serializers.AltadesistenciaSerializer



class SelecaoListCreateAPIView(generics.ListCreateAPIView):
    permission_classes = (IsAuthenticated, GlobalDefaultPermission,)
    queryset = models.Selecao.objects.select_related(
        'fk_terapeuta_avaliador__fk_associado',
        'fk_associado_avaliado'
    )
    serializer_class = serializers.SelecaoSerializer
    
    def get_queryset(self):
        """Filtros opcionais via query params"""
        queryset = super().get_queryset()
        
        # Filtro por avaliador
        avaliador_id = self.request.query_params.get('avaliador', None)
        if avaliador_id:
            queryset = queryset.filter(fk_terapeuta_avaliador_id=avaliador_id)
            
        # Filtro por avaliado
        avaliado_id = self.request.query_params.get('avaliado', None)
        if avaliado_id:
            queryset = queryset.filter(fk_associado_avaliado_id=avaliado_id)
            
        # Filtro por data
        data_inicio = self.request.query_params.get('data_inicio', None)
        data_fim = self.request.query_params.get('data_fim', None)
        
        if data_inicio:
            queryset = queryset.filter(dat_avaliacao__gte=data_inicio)
        if data_fim:
            queryset = queryset.filter(dat_avaliacao__lte=data_fim)
            
        return queryset.order_by('-dat_avaliacao')



class SelecaoRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = (IsAuthenticated, GlobalDefaultPermission,)
    queryset = models.Selecao.objects.select_related(
        'fk_terapeuta_avaliador__fk_associado',
        'fk_associado_avaliado'
    )
    serializer_class = serializers.SelecaoSerializer


